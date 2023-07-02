#!/usr/bin/python

import random, time, statistics

from openpyxl import load_workbook

from termcolor import colored, cprint

w = load_workbook("perf.xlsx")

s = w.active

l = ["0", "1", "2", "3", "4", "5", "6", "7", "8", "9"]

lb = ["0", "1", "2", "3"]

easy_l = [1, 2, 3]

medium_l = [4, 5]

hard_l = [6, 7, 8]

easy_lb = [1]

medium_lb = [1 ,2]

hard_lb = [3, 4]

cl = ["+", "-"]

cl3 = ["c", "%"]

cache_l = []

result_str = ""

many = int(input("How many calculates ? "))

points = 0

difficulty = str(input("Difficlty (easy, medium, hard)"))

perf1_l = []

perf2_l = []

for i in range (0, many):

    if difficulty == "easy":

        long = random.choice(easy_l)

    if difficulty == "medium":

        long = random.choice(medium_l)

    if difficulty == "hard":

        long = random.choice(hard_l)

    pc = random.choice(cl3)

    if pc == "c":

        for i2 in range (0, 2):

            for i3 in range (0, long):

                r = random.choice(l)

                result_str = result_str + r

            cache_l.append(result_str)

            result_str = ""

        appl = random.choice(cl)

        if appl == "-":

            veritas = int(cache_l[-2]) - int(cache_l[-1])

        else:

            veritas = int(cache_l[-2]) + int(cache_l[-1])

        phrase = "What's " + str(cache_l[-2]) + " " + appl + " " + str(cache_l[-1]) + "? "

        strt = time.time()

        veritas_h = str(input(phrase))

        veritas_h = int(veritas_h)

        end = time.time()

        phrase = "time: " + str(end - strt) + " s"

        if veritas == veritas_h:

            perf1_l.append("correct")

            cprint("correct", "green")

            cprint(phrase, "cyan")

        else:

            perf1_l.append("wrong")

            cprint("wrong", "red")

            cprint(phrase, "cyan")

        perf2_l.append(end - strt)

        phrase = "result:" + str(veritas)

        cprint(phrase, "green")

    if pc == "%":

        if difficulty == "easy":

            long = random.choice(easy_lb)

        if difficulty == "medium":

            long = random.choice(medium_lb)

        if difficulty == "hard":

            long = random.choice(hard_lb)

        for i3 in range (0, long):

            r = random.choice(l)

            result_str = result_str + r

        cache_l.append(result_str)

        appl = random.choice(cl)

        prct = random.choice(lb) + "." + str(random.choice(l))

        cache_l.append(prct)

        strt = time.time()

        veritas = int(cache_l[-2]) * float(cache_l[-1])

        phrase = "What's " + str(cache_l[-2]) + " * " + str(cache_l[-1]) + "? "

        veritas_h = str(input(phrase))

        veritas_h = float(veritas_h)

        end = time.time()

        phrase = "time: " + str(end - strt) + " s"

        if veritas == veritas_h:

            perf1_l.append("correct")

            cprint("correct", "green")

            cprint(phrase, "cyan")

        else:

            perf1_l.append("wrong")

            cprint("wrong", "red")

            cprint(phrase, "cyan")

        perf2_l.append(end - strt)

        phrase = "result:" + str(veritas)

        cprint(phrase, "green")

t = 1

while s.cell(row=t, column=1).value != None:

    t += 1

s.cell(row= t, column=1).value = perf1_l.count("correct") 

s.cell(row= t, column=2).value = perf1_l.count("wrong") 

s.cell(row= t, column=3).value = statistics.mean(perf2_l) 

s.cell(row= t, column=4).value = many 

w.save("perf.xlsx")





